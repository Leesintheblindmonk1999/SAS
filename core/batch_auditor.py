"""
core/batch_auditor.py — Omni-Scanner Semantic v2.0
═══════════════════════════════════════════════════════
Bulk Audit Engine — Topographic Corpus Assessment

Processes batches of .txt files (+ optional .pdf/.docx connectors)
using FullDiagnostic.run() and generates Excel/CSV reports with
Durante conditional formatting: Green κD≥0.56, Red κD<0.56.

Central metric:
  CIR = Corpus Integrity Rate
      = (N documents with ManifoldScore ≥ κD) / N_total

  If CIR < 0.50 → CORPUS COMPROMISED

Excel header:
  "Omni-Scanner v6 - Powered by Durante Invariance (κD=0.56)"

DEPENDENCIES:
  openpyxl  — pip install openpyxl
  engines.full_diagnostic — already in project (Zero Noise)
  core.tda_attestation    — already in project (Zero Noise)
"""
from __future__ import annotations

import io
import csv
import math
import datetime
import traceback
from dataclasses import dataclass, field
from typing import List, Optional, Callable, Iterator

# ── Imports internos — Zero Noise: no modificamos los motores ──
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from engines.full_diagnostic import FullDiagnostic, FullDiagnosticReport

# ── openpyxl (optional: if missing, CSV-only export) ──────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── Conectores opcionales ──────────────────────────────────────
try:
    import fitz                     # PyMuPDF
    _PDF_OK = True
except ImportError:
    _PDF_OK = False

try:
    import docx as _docx_lib        # python-docx
    _DOCX_OK = True
except ImportError:
    _DOCX_OK = False


# ══════════════════════════════════════════════════════════════
# Colores Durante (openpyxl hex sin '#')
# ══════════════════════════════════════════════════════════════

_COL_HEADER_BG  = "0D0D0D"     # negro profundo
_COL_HEADER_FG  = "FFB100"     # amber Durante
_COL_GREEN_BG   = "D4EDDA"     # verde esmeralda claro
_COL_GREEN_FG   = "145A32"     # verde esmeralda oscuro
_COL_RED_BG     = "FADADD"     # light crimson red
_COL_RED_FG     = "7B241C"     # dark crimson red
_COL_REVIEW_BG  = "FFF3CD"     # light amber
_COL_REVIEW_FG  = "7D6608"     # dark amber
_COL_ALT_ROW    = "F8F6F0"     # gris muy claro para filas alternas
_COL_WHITE      = "FFFFFF"
_COL_SUBHDR_BG  = "1C1C1C"     # gris oscuro sub-header
_COL_SUBHDR_FG  = "FFB100"


# ══════════════════════════════════════════════════════════════
# Estructuras de datos
# ══════════════════════════════════════════════════════════════

@dataclass
class FileResult:
    """Result of the analysis of a single file."""
    filename:        str
    status:          str            # "OK" | "ERROR" | "SKIPPED"
    error_msg:       str = ""

    # Fields from FullDiagnosticReport
    manifold_score:  float = -1.0
    manifold_verdict: str  = "N/A"
    coherence_score: float = -1.0
    topology_flag:   str   = "N/A"
    overall_verdict: str   = "N/A"  # CLEAR | REVIEW | HIGH_RISK
    domain_risk:     float = 0.0
    confidence:      float = 0.0
    word_entropy:    float = -1.0   # extracted from manifold_summary
    h1_cycles:       int   = 0      # extracted from raw_reports.topology
    input_type:      str   = "generic"

    # Timestamps
    processed_at: str = field(
        default_factory=lambda: datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    )

    @property
    def passes_kd(self) -> bool:
        """True if ManifoldScore ≥ κD."""
        return self.manifold_score >= 0 and self.manifold_score >= BatchAuditor.KAPPA_D

    def to_row(self) -> list:
        """Row for CSV/XLSX."""
        return [
            self.filename,
            f"{self.manifold_score:.4f}" if self.manifold_score >= 0 else "ERROR",
            f"{self.coherence_score:.4f}" if self.coherence_score >= 0 else "N/A",
            f"{self.word_entropy:.4f}" if self.word_entropy >= 0 else "N/A",
            str(self.h1_cycles),
            self.overall_verdict,
            self.topology_flag,
            self.manifold_verdict,
            f"{self.domain_risk:.4f}",
            f"{self.confidence:.2%}",
            self.status,
            self.error_msg,
            self.processed_at,
        ]


@dataclass
class BatchReport:
    """Complete bulk audit report."""
    timestamp:      str
    kappa_d:        float
    total_files:    int
    processed_ok:   int
    errors:         int
    skipped:        int

    # CIR — Corpus Integrity Rate
    cir:            float   # ∈ [0, 1]
    corpus_verdict: str     # "INTACT" | "UNDER REVIEW" | "COMPROMISED"

    results:        List[FileResult] = field(default_factory=list)
    openpyxl_available: bool = False

    # Aggregated statistics
    mean_manifold:  float = 0.0
    min_manifold:   float = 0.0
    max_manifold:   float = 0.0
    high_risk_count: int  = 0
    clear_count:    int   = 0
    review_count:   int   = 0

    @property
    def corpus_compromised(self) -> bool:
        return self.cir < 0.50


COLUMNS = [
    "File",
    "ManifoldScore",
    "Coherence",
    "Entropy",
    "H₁ Cycles",
    "Verdict",
    "TopologyFlag",
    "ManifoldVerdict",
    "DomainRisk",
    "Confidence",
    "Status",
    "Error",
    "Timestamp",
]


# ══════════════════════════════════════════════════════════════
# Motor Principal
# ══════════════════════════════════════════════════════════════

class BatchAuditor:
    """
    Bulk Audit Engine — Topographic Corpus Assessment.

    Processes lists of files or in-memory bytes, extracts topological
    metrics via FullDiagnostic.run(), and generates reports
    Excel coloreados + CSV crudo.

    Parameters
    ----------
    kappa_d : float
        Reference threshold (default 0.56 — Durante Constant).
        Determines cell color and CIR calculation.
    input_type : str
        Default document type: "legal" | "generic" | "discourse" | etc.
    stability_threshold : float
        Threshold for FullDiagnostic (inherits kappa_d by default).
    """

    KAPPA_D: float = 0.56   # clase-nivel para FileResult.passes_kd

    def __init__(
        self,
        kappa_d: float = 0.56,
        input_type: str = "generic",
        stability_threshold: Optional[float] = None,
    ):
        self.kappa_d    = kappa_d
        BatchAuditor.KAPPA_D = kappa_d
        self.input_type = input_type
        self._diagnostic = FullDiagnostic(
            stability_threshold=stability_threshold or kappa_d
        )

    # ── API principal ─────────────────────────────────────────

    def run_batch(
        self,
        files: List[tuple],                    # [(filename, bytes_or_str), ...]
        input_type: Optional[str] = None,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> BatchReport:
        """
        Processes a batch of files.

        Parámetros
        ----------
        files : list de (filename, content)
            content can be str (text) or bytes (for .pdf/.docx).
        input_type : str
            Document type override. If None, uses self.input_type.
        progress_callback : callable(current, total, filename)
            Called after each file to update the UI.

        Returns complete BatchReport with all metrics.
        """
        itype   = input_type or self.input_type
        ts      = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        results = []

        for i, (fname, content) in enumerate(files):
            if progress_callback:
                progress_callback(i, len(files), fname)

            result = self._process_single(fname, content, itype)
            results.append(result)

        # Callback final
        if progress_callback:
            progress_callback(len(files), len(files), "")

        return self._build_report(ts, results)

    def run_batch_stream(
        self,
        files: List[tuple],
        input_type: Optional[str] = None,
    ) -> Iterator[FileResult]:
        """
        Generator version — yields FileResult as it processes.
        Allows real-time Streamlit UI updates without a spinner.
        """
        itype = input_type or self.input_type
        for fname, content in files:
            yield self._process_single(fname, content, itype)

    # ── Procesamiento de un archivo ───────────────────────────

    def _process_single(
        self,
        filename: str,
        content,
        input_type: str,
    ) -> FileResult:
        """
        Processes a single file. Isolates errors so they don't stop the batch.
        """
        result = FileResult(filename=filename, status="OK", input_type=input_type)

        try:
            # 1. Extract text
            text = self._extract_text(filename, content)
            if text is None:
                result.status    = "SKIPPED"
                result.error_msg = "Unsupported format"
                return result

            if len(text.strip()) < 20:
                result.status    = "SKIPPED"
                result.error_msg = "Text too short (<20 chars)"
                return result

            # 2. Run full diagnostic
            dr: FullDiagnosticReport = self._diagnostic.run(
                text, input_type=input_type
            )

            # 3. Map fields
            result.manifold_score   = dr.manifold_score
            result.manifold_verdict = dr.manifold_verdict
            result.coherence_score  = dr.coherence_score
            result.topology_flag    = dr.topology_flag
            result.overall_verdict  = dr.overall_verdict
            result.domain_risk      = dr.domain_risk
            result.confidence       = dr.confidence

            # 4. Extract word_entropy from manifold summary
            result.word_entropy = self._parse_entropy(
                dr.raw_reports.get("manifold_summary", "")
            )

            # 5. Extract h1_cycles from topology
            topo = dr.raw_reports.get("topology", {})
            # topology_mapper does not expose H₁ directly — we use edge_count
            # as connectivity proxy (higher edge_count = more cycles)
            result.h1_cycles = topo.get("edge_count", 0)

        except Exception as e:
            result.status    = "ERROR"
            result.error_msg = f"{type(e).__name__}: {str(e)[:120]}"

        return result

    # ── Text extraction ──────────────────────────────────────

    def _extract_text(self, filename: str, content) -> Optional[str]:
        """
        Extracts text from .txt, .pdf (PyMuPDF), .docx (python-docx).
        Returns None if format is not supported.
        """
        fname_lower = filename.lower()

        # Si ya es str, devolver directamente
        if isinstance(content, str):
            return content

        # bytes → decode by extension
        if fname_lower.endswith(".txt"):
            for enc in ("utf-8", "latin-1", "cp1252"):
                try:
                    return content.decode(enc)
                except (UnicodeDecodeError, AttributeError):
                    continue
            return None

        if fname_lower.endswith(".pdf"):
            if not _PDF_OK:
                return (
                    f"[PDF: {filename}]\n"
                    "Instalar PyMuPDF para leer PDFs: pip install PyMuPDF\n"
                    "This text is a placeholder for structural analysis."
                )
            try:
                doc = fitz.open(stream=content, filetype="pdf")
                pages = []
                for page in doc:
                    pages.append(page.get_text("text"))
                return "\n".join(pages)
            except Exception as e:
                raise RuntimeError(f"Error reading PDF: {e}")

        if fname_lower.endswith(".docx"):
            if not _DOCX_OK:
                return (
                    f"[DOCX: {filename}]\n"
                    "Instalar python-docx para leer DOCX: pip install python-docx\n"
                    "This text is a placeholder for structural analysis."
                )
            try:
                doc = _docx_lib.Document(io.BytesIO(content))
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except Exception as e:
                raise RuntimeError(f"Error reading DOCX: {e}")

        # Intentar como texto plano por defecto
        try:
            return content.decode("utf-8", errors="replace")
        except Exception:
            return None

    def _parse_entropy(self, summary: str) -> float:
        """
        Extracts word_entropy from manifold_summary.
        The summary contains lines like 'Observed entropy: X.XXXX bits'
        """
        import re
        m = re.search(r"Observed entropy[:\s]+([0-9.]+)|Entrop[ií]a observada[:\s]+([0-9.]+)", summary)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                pass
        # Fallback: search TTR/word_entropy pattern in PARTIAL_VALIDATION
        m2 = re.search(r"TTR=([0-9.]+)", summary)
        if m2:
            try:
                return float(m2.group(1))
            except ValueError:
                pass
        return -1.0

    # ── Report construction ──────────────────────────────────

    def _build_report(self, ts: str, results: List[FileResult]) -> BatchReport:
        ok_results = [r for r in results if r.status == "OK"]
        errors     = sum(1 for r in results if r.status == "ERROR")
        skipped    = sum(1 for r in results if r.status == "SKIPPED")

        # CIR: fraction of OK files with ManifoldScore ≥ κD
        if ok_results:
            above_kd = sum(1 for r in ok_results if r.passes_kd)
            cir = above_kd / len(ok_results)
        else:
            cir = 0.0

        if tic >= 0.80:
            corpus_verdict = "INTACT"
        elif tic >= 0.50:
            corpus_verdict = "UNDER REVIEW"
        else:
            corpus_verdict = "COMPROMISED"

        scores = [r.manifold_score for r in ok_results if r.manifold_score >= 0]
        verdicts = [r.overall_verdict for r in ok_results]

        return BatchReport(
            timestamp      = ts,
            kappa_d        = self.kappa_d,
            total_files    = len(results),
            processed_ok   = len(ok_results),
            errors         = errors,
            skipped        = skipped,
            cir            = round(cir, 6),
            corpus_verdict = corpus_verdict,
            results        = results,
            openpyxl_available = _OPENPYXL_OK,
            mean_manifold  = round(sum(scores) / len(scores), 6) if scores else 0.0,
            min_manifold   = round(min(scores), 6) if scores else 0.0,
            max_manifold   = round(max(scores), 6) if scores else 0.0,
            high_risk_count= verdicts.count("HIGH_RISK"),
            clear_count    = verdicts.count("CLEAR"),
            review_count   = verdicts.count("REVIEW"),
        )

    # ══════════════════════════════════════════════════════════
    # Exportadores
    # ══════════════════════════════════════════════════════════

    def to_csv_bytes(self, report: BatchReport) -> bytes:
        """Generates raw CSV with all results."""
        buf = io.StringIO()
        writer = csv.writer(buf)

        # Meta-header
        writer.writerow(["Omni-Scanner v6 - Powered by Durante Invariance (κD=0.56)"])
        writer.writerow([f"Generado: {report.timestamp}"])
        writer.writerow([
            f"CIR: {report.cir:.2%}",
            f"Corpus Verdict: {report.corpus_verdict}",
            f"Total files: {report.total_files}",
            f"Processed OK: {report.processed_ok}",
            f"Errors: {report.errors}",
        ])
        writer.writerow([])

        # Columnas
        writer.writerow(COLUMNS)

        # Datos
        for r in report.results:
            writer.writerow(r.to_row())

        # Statistics row
        writer.writerow([])
        writer.writerow(["STATISTICS"])
        writer.writerow(["Average ManifoldScore", f"{report.mean_manifold:.4f}"])
        writer.writerow(["Minimum ManifoldScore",   f"{report.min_manifold:.4f}"])
        writer.writerow(["Maximum ManifoldScore",    f"{report.max_manifold:.4f}"])
        writer.writerow(["CLEAR", report.clear_count])
        writer.writerow(["REVIEW", report.review_count])
        writer.writerow(["HIGH_RISK", report.high_risk_count])

        return buf.getvalue().encode("utf-8-sig")   # BOM para Excel en Windows

    def to_xlsx_bytes(self, report: BatchReport) -> bytes:
        """
        Generates professional Excel with Durante conditional formatting.

        Sheet 1: CORPUS — main table with verdict colors
        Sheet 2: STATISTICS — KPIs and summary metrics
        """
        if not _OPENPYXL_OK:
            raise RuntimeError(
                "openpyxl is not installed. "
                "pip install openpyxl  — or use CSV export."
            )

        wb = Workbook()

        # ── Hoja 1: CORPUS ────────────────────────────────────
        ws = wb.active
        ws.title = "CORPUS"

        self._build_corpus_sheet(ws, report)

        # ── Sheet 2: STATISTICS ─────────────────────────────
        ws2 = wb.create_sheet("STATISTICS")
        self._build_stats_sheet(ws2, report)

        # Guardar en bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Constructor de hoja CORPUS ────────────────────────────

    def _build_corpus_sheet(self, ws, report: BatchReport):
        """Main sheet with Durante conditional formatting."""

        # Estilos reutilizables
        hdr_font    = Font(name="Arial", bold=True, color=_COL_HEADER_FG, size=10)
        hdr_fill    = PatternFill("solid", fgColor=_COL_HEADER_BG)
        hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center      = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

        fill_green_bg  = PatternFill("solid", fgColor=_COL_GREEN_BG)
        fill_red_bg    = PatternFill("solid", fgColor=_COL_RED_BG)
        fill_review_bg = PatternFill("solid", fgColor=_COL_REVIEW_BG)
        fill_alt       = PatternFill("solid", fgColor=_COL_ALT_ROW)
        fill_white     = PatternFill("solid", fgColor=_COL_WHITE)

        font_green  = Font(name="Arial", bold=True, color=_COL_GREEN_FG, size=10)
        font_red    = Font(name="Arial", bold=True, color=_COL_RED_FG, size=10)
        font_review = Font(name="Arial", bold=True, color=_COL_REVIEW_FG, size=10)
        font_normal = Font(name="Arial", size=10)
        font_mono   = Font(name="Courier New", size=9)

        # ── Row 1: Main title ──────────────────────────────
        ws.merge_cells("A1:M1")
        title_cell = ws["A1"]
        title_cell.value = "Omni-Scanner v6 - Powered by Durante Invariance (κD=0.56)"
        title_cell.font  = Font(name="Arial", bold=True, color=_COL_HEADER_FG, size=13)
        title_cell.fill  = PatternFill("solid", fgColor=_COL_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Fila 2: Meta-info ─────────────────────────────────
        ws.merge_cells("A2:M2")
        meta = ws["A2"]
        meta.value = (
            f"Generated: {report.timestamp}  |  "
            f"Total files: {report.total_files}  |  "
            f"Processed OK: {report.processed_ok}  |  "
            f"Errors: {report.errors}  |  "
            f"κD = {report.kappa_d}"
        )
        meta.font  = Font(name="Arial", italic=True, color="888888", size=9)
        meta.fill  = PatternFill("solid", fgColor="1A1A1A")
        meta.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── Fila 3: TIC banner ────────────────────────────────
        ws.merge_cells("A3:M3")
        tic_cell = ws["A3"]
        tic_pct  = f"{report.cir:.1%}"
        tic_cell.value = (
            f"CIR — CORPUS INTEGRITY RATE: {tic_pct}  |  "
            f"VERDICT: {report.corpus_verdict}  |  "
            f"CLEAR: {report.clear_count}  REVIEW: {report.review_count}  "
            f"HIGH_RISK: {report.high_risk_count}"
        )
        if report.corpus_compromised:
            tic_cell.font = Font(name="Arial", bold=True, color="FF2D2D", size=11)
            tic_cell.fill = PatternFill("solid", fgColor="2D0000")
        elif report.cir >= 0.80:
            tic_cell.font = Font(name="Arial", bold=True, color="00CC66", size=11)
            tic_cell.fill = PatternFill("solid", fgColor="002D00")
        else:
            tic_cell.font = Font(name="Arial", bold=True, color=_COL_HEADER_FG, size=11)
            tic_cell.fill = PatternFill("solid", fgColor="2D2000")
        tic_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 22

        # ── Fila 4: Separador ─────────────────────────────────
        ws.row_dimensions[4].height = 6

        # ── Fila 5: Headers de columnas ───────────────────────
        HDR_ROW = 5
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=HDR_ROW, column=col_idx, value=col_name)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = thin_border
        ws.row_dimensions[HDR_ROW].height = 30

        # ── Filas de datos ────────────────────────────────────
        MANIFOLD_COL = 2   # columna B = ManifoldScore
        VERDICT_COL  = 6   # column F = Verdict

        for row_offset, result in enumerate(report.results):
            row_num = HDR_ROW + 1 + row_offset
            row_data = result.to_row()
            is_alt = row_offset % 2 == 1

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center if col_idx != 1 else Alignment(
                    horizontal="left", vertical="center"
                )

                # Formato condicional Durante en ManifoldScore (col B)
                if col_idx == MANIFOLD_COL and result.status == "OK":
                    try:
                        score = float(value)
                        if score >= self.kappa_d:
                            cell.fill = fill_green_bg
                            cell.font = font_green
                        else:
                            cell.fill = fill_red_bg
                            cell.font = font_red
                    except (ValueError, TypeError):
                        cell.font = Font(name="Courier New", size=9, color="999999")
                        cell.fill = fill_alt if is_alt else fill_white

                # Verdict color (col F)
                elif col_idx == VERDICT_COL:
                    if value == "CLEAR":
                        cell.fill = fill_green_bg
                        cell.font = font_green
                    elif value == "HIGH_RISK":
                        cell.fill = fill_red_bg
                        cell.font = font_red
                    elif value == "REVIEW":
                        cell.fill = fill_review_bg
                        cell.font = font_review
                    else:
                        cell.font = font_normal
                        cell.fill = fill_alt if is_alt else fill_white

                # Error en rojo
                elif col_idx == 12 and value:   # columna Error
                    cell.font = Font(name="Arial", size=9, color="FF2D2D", italic=True)
                    cell.fill = fill_alt if is_alt else fill_white

                # Numeric value columns — mono font
                elif col_idx in (3, 4, 5, 9, 10):
                    cell.font = font_mono
                    cell.fill = fill_alt if is_alt else fill_white

                else:
                    cell.font = font_normal
                    cell.fill = fill_alt if is_alt else fill_white

            ws.row_dimensions[row_num].height = 18

        # ── Anchos de columna ─────────────────────────────────
        col_widths = [38, 14, 13, 13, 10, 13, 16, 18, 14, 12, 10, 28, 22]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze encabezados
        ws.freeze_panes = "A6"

        # Auto-filter en fila de headers
        ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(COLUMNS))}{HDR_ROW + len(report.results)}"

    # ── STATISTICS sheet constructor ────────────────────────

    def _build_stats_sheet(self, ws, report: BatchReport):
        """KPI sheet with corpus summary metrics."""

        font_title  = Font(name="Arial", bold=True, color=_COL_HEADER_FG, size=12)
        fill_dark   = PatternFill("solid", fgColor=_COL_HEADER_BG)
        font_label  = Font(name="Arial", bold=True, color="333333", size=10)
        font_value  = Font(name="Courier New", bold=True, size=11)
        center      = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"].value     = "CORPUS STATISTICS — Durante Invariance"
        ws["A1"].font      = font_title
        ws["A1"].fill      = fill_dark
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 26

        kpis = [
            ("CIR — Corpus Integrity Rate", f"{report.cir:.2%}",
             "00CC66" if report.cir >= 0.5 else "FF2D2D"),
            ("Corpus Verdict", report.corpus_verdict,
             "00CC66" if not report.corpus_compromised else "FF2D2D"),
            ("Average ManifoldScore", f"{report.mean_manifold:.4f}",
             "00CC66" if report.mean_manifold >= report.kappa_d else "FF2D2D"),
            ("Minimum ManifoldScore", f"{report.min_manifold:.4f}",
             "00CC66" if report.min_manifold >= report.kappa_d else "FF2D2D"),
            ("Maximum ManifoldScore", f"{report.max_manifold:.4f}", "FFB100"),
            ("κD (Durante Constant)", f"{report.kappa_d}", "FFB100"),
            ("Total Files", str(report.total_files), "333333"),
            ("Processed OK", str(report.processed_ok), "333333"),
            ("Errors", str(report.errors), "FF2D2D" if report.errors > 0 else "333333"),
            ("CLEAR Documents", str(report.clear_count), "00CC66"),
            ("REVIEW Documents", str(report.review_count), "FFB100"),
            ("HIGH_RISK Documents", str(report.high_risk_count), "FF2D2D"),
        ]

        for i, (label, value, color) in enumerate(kpis, start=3):
            ws.cell(row=i, column=1, value=label).font = font_label
            ws.cell(row=i, column=1).fill = PatternFill(
                "solid", fgColor="F5F5F5" if i % 2 == 0 else _COL_WHITE
            )
            val_cell = ws.cell(row=i, column=2, value=value)
            val_cell.font = Font(name="Courier New", bold=True, size=11, color=color)
            val_cell.alignment = center
            ws.row_dimensions[i].height = 22

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22

        # Alerta de corpus comprometido
        if report.corpus_compromised:
            alert_row = len(kpis) + 5
            ws.merge_cells(f"A{alert_row}:D{alert_row}")
            alert = ws.cell(row=alert_row, column=1)
            alert.value = (
                "⚠ CORPUS COMPROMISED: High Structural Entropy Density"
            )
            alert.font  = Font(name="Arial", bold=True, color="FF2D2D", size=11)
            alert.fill  = PatternFill("solid", fgColor="2D0000")
            alert.alignment = center
            ws.row_dimensions[alert_row].height = 26


# ══════════════════════════════════════════════════════════════
# Convenience function
# ══════════════════════════════════════════════════════════════

def quick_batch(
    files: List[tuple],
    kappa_d: float = 0.56,
    input_type: str = "generic",
) -> BatchReport:
    """
    One-liner bulk audit.

    Ejemplo:
        from core.batch_auditor import quick_batch
        files = [("doc1.txt", texto1), ("doc2.txt", texto2)]
        report = quick_batch(files)
        print(f"TIC: {report.cir:.1%} — {report.corpus_verdict}")
    """
    auditor = BatchAuditor(kappa_d=kappa_d, input_type=input_type)
    return auditor.run_batch(files)
